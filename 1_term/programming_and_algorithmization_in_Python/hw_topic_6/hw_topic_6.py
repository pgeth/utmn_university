# Задание
# Для каждого user_id в файле visit_log.csv определите третий столбец с категорией покупки. Если покупка была, сам файл visit_log.csv изменять не надо.
# Запишите в файл funnel.csv визиты из файла visit_log.csv, в которых были покупки с указанием категории.

# Учтите условие для данных:

# содержимое visit_log.csv — не помещается в оперативную память компьютера, используйте только построчную обработку этого файла.
# Примеры работы программы при выводе первых трёх строк файла funnel.csv:
# user_id,source,category
# 1840e0b9d4,other,Продукты
# 4e4f90fcfb,context,Электроника

import csv
import json
from openpyxl import load_workbook

def read_purchases():
    purchases = {}
    with open('purchase_log.txt', 'r', encoding='utf-8') as f:
        next(f)
        for row in f:
            info = json.loads(row)
            purchases[info['user_id']] = info['category']
    return purchases

def create_funnel(purchases):
    wb = load_workbook('visit_log.xlsx', read_only=True)
    ws = wb.active

    with open('funnel.csv', 'w', encoding='utf-8', newline='') as f_out:
        writer = csv.writer(f_out)
        writer.writerow(['user_id', 'source', 'category'])

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  
                continue

            user_id = row[0]  
            source = row[1]  

            if user_id in purchases:
                writer.writerow([user_id, source, purchases[user_id]])

    wb.close()

purchases = read_purchases()
create_funnel(purchases)
